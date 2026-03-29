import shutil
from pathlib import Path
from typing import cast

import openpyxl
import pytest

from rfexcel.exception.library_exceptions import (
    HeadersNotDeterminedException, NullComponentException,
    RowIndexOutOfBoundsException, WorkbookNotOpenException)
from rfexcel.RFExcelLibrary import RFExcelLibrary
from tests.pyth.conftest import CSV_FILE, XLS_FILE, XLSX_FILE

_HEADERS = ["Product ID", "Description", "Price", "Location"]
_NEW_ROW = {"Product ID": "P-NEW", "Description": "Inserted", "Price": 5.55, "Location": "Depot"}


# ---------------------------------------------------------------------------
# XLSX – Edit mode
# ---------------------------------------------------------------------------

class TestInsertRowXlsxEdit:

    def test_row_inserted_at_correct_position(self, lib: RFExcelLibrary, tmp_path: Path):
        path = str(shutil.copy(XLSX_FILE, tmp_path / "data.xlsx"))
        lib.load_workbook(path)
        rows_before = lib.get_rows()
        first_row_before = cast(str, rows_before[0]["Product ID"])

        lib.insert_row(_NEW_ROW, row=2)
        rows_after = lib.get_rows()

        assert len(rows_after) == len(rows_before) + 1
        assert rows_after[0]["Product ID"] == "P-NEW"
        assert rows_after[0]["Description"] == "Inserted"
        assert rows_after[0]["Price"] == 5.55
        assert rows_after[0]["Location"] == "Depot"
        assert rows_after[1]["Product ID"] == first_row_before

    def test_partial_row_fills_missing_columns_with_empty_string(
        self, lib: RFExcelLibrary, tmp_path: Path
    ):
        path = str(shutil.copy(XLSX_FILE, tmp_path / "data.xlsx"))
        lib.load_workbook(path)
        lib.insert_row({"Product ID": "P-PARTIAL"}, row=2)
        inserted = lib.get_rows()[0]
        assert inserted["Product ID"] == "P-PARTIAL"
        assert inserted["Description"] == ""
        assert inserted["Price"] == ""
        assert inserted["Location"] == ""

    def test_unknown_keys_are_silently_ignored(self, lib: RFExcelLibrary, tmp_path: Path):
        path = str(shutil.copy(XLSX_FILE, tmp_path / "data.xlsx"))
        lib.load_workbook(path)
        before = len(lib.get_rows())
        lib.insert_row({"Product ID": "P-777", "NonExistent": "ignored"}, row=2)
        rows = lib.get_rows()
        assert len(rows) == before + 1
        assert rows[0]["Product ID"] == "P-777"

    def test_row_is_persisted_after_save(self, lib: RFExcelLibrary, tmp_path: Path):
        path = str(shutil.copy(XLSX_FILE, tmp_path / "data.xlsx"))
        lib.load_workbook(path)
        lib.insert_row(_NEW_ROW, row=2)
        lib.save_workbook()
        lib.close()

        lib2 = RFExcelLibrary()
        lib2.load_workbook(path)
        rows = lib2.get_rows()
        assert rows[0]["Product ID"] == "P-NEW"
        lib2.close()

    def test_insert_at_last_data_row(self, lib: RFExcelLibrary, tmp_path: Path):
        path = str(shutil.copy(XLSX_FILE, tmp_path / "data.xlsx"))
        lib.load_workbook(path)
        rows_before = lib.get_rows()
        last_before = cast(str, rows_before[-1]["Product ID"])
        insert_at = len(rows_before) + 1

        lib.insert_row({"Product ID": "P-LAST"}, row=insert_at)
        rows_after = lib.get_rows()

        assert rows_after[-2]["Product ID"] == "P-LAST"
        assert rows_after[-1]["Product ID"] == last_before

    def test_custom_header_row(self, lib: RFExcelLibrary, tmp_path: Path):
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["filler"])
        ws.append(["Name", "Score"])
        ws.append(["Alice", 90])
        ws.append(["Bob", 80])
        out = str(tmp_path / "custom.xlsx")
        wb.save(out)

        lib.load_workbook(out)
        lib.insert_row({"Name": "Charlie", "Score": 95}, row=3, header_row=2)
        rows = lib.get_rows(header_row=2)
        assert rows[0]["Name"] == "Charlie"
        assert rows[0]["Score"] == 95
        assert rows[1]["Name"] == "Alice"

    def test_row_equal_to_header_row_raises(self, lib: RFExcelLibrary, tmp_path: Path):
        path = str(shutil.copy(XLSX_FILE, tmp_path / "data.xlsx"))
        lib.load_workbook(path)
        with pytest.raises(RowIndexOutOfBoundsException):
            lib.insert_row(_NEW_ROW, row=1, header_row=1)

    def test_row_less_than_header_row_raises(self, lib: RFExcelLibrary, tmp_path: Path):
        path = str(shutil.copy(XLSX_FILE, tmp_path / "data.xlsx"))
        lib.load_workbook(path)
        with pytest.raises(RowIndexOutOfBoundsException):
            lib.insert_row(_NEW_ROW, row=1, header_row=2)

    def test_header_row_out_of_range_raises(self, lib: RFExcelLibrary, tmp_path: Path):
        path = str(shutil.copy(XLSX_FILE, tmp_path / "data.xlsx"))
        lib.load_workbook(path)
        with pytest.raises(HeadersNotDeterminedException):
            lib.insert_row(_NEW_ROW, row=9999, header_row=9998)


# ---------------------------------------------------------------------------
# Read-only / streaming modes – raises for all formats
# ---------------------------------------------------------------------------

@pytest.mark.parametrize(
    "path",
    [XLSX_FILE, XLS_FILE, CSV_FILE],
    ids=["xlsx_stream", "xls_on_demand", "csv_stream"],
)
def test_insert_row_raises_in_read_only_mode(lib: RFExcelLibrary, path: str):
    lib.load_workbook(path, read_only=True)
    with pytest.raises(NullComponentException):
        lib.insert_row(_NEW_ROW, row=2)


# ---------------------------------------------------------------------------
# XLS – Edit mode
# ---------------------------------------------------------------------------

class TestInsertRowXlsEdit:

    def test_insert_row_triggers_conversion_and_persists(
        self, lib: RFExcelLibrary, tmp_path: Path
    ):
        path = str(shutil.copy(XLS_FILE, tmp_path / "example.xls"))
        new_path = str(tmp_path / "result.xlsx")
        lib.load_workbook(path)
        lib.insert_row({"First Name": "Jane", "Last Name": "Doe"}, row=2)
        lib.save_workbook(new_path)
        lib.close()

        lib2 = RFExcelLibrary()
        lib2.load_workbook(new_path)
        rows = lib2.get_rows()
        assert rows[0]["First Name"] == "Jane"
        assert rows[0]["Last Name"] == "Doe"
        lib2.close()

    def test_original_xls_untouched_after_insert_row(
        self, lib: RFExcelLibrary, tmp_path: Path
    ):
        path = str(shutil.copy(XLS_FILE, tmp_path / "example.xls"))
        original_lib = RFExcelLibrary()
        original_lib.load_workbook(path)
        before = len(original_lib.get_rows())
        original_lib.close()

        lib.load_workbook(path)
        lib.insert_row(_NEW_ROW, row=2)
        lib.save_workbook(str(tmp_path / "out.xlsx"))
        lib.close()

        lib2 = RFExcelLibrary()
        lib2.load_workbook(path)
        assert len(lib2.get_rows()) == before
        lib2.close()


# ---------------------------------------------------------------------------
# CSV – Edit mode
# ---------------------------------------------------------------------------

class TestInsertRowCsvEdit:

    def test_row_inserted_at_correct_position(self, lib: RFExcelLibrary, tmp_path: Path):
        path = str(shutil.copy(CSV_FILE, tmp_path / "data.csv"))
        lib.load_workbook(path)
        rows_before = lib.get_rows()
        first_before = cast(str, rows_before[0]["Product ID"])

        lib.insert_row(_NEW_ROW, row=2)
        rows_after = lib.get_rows()

        assert len(rows_after) == len(rows_before) + 1
        assert rows_after[0]["Product ID"] == "P-NEW"
        assert rows_after[1]["Product ID"] == first_before

    def test_partial_row_fills_missing_with_empty_string(
        self, lib: RFExcelLibrary, tmp_path: Path
    ):
        path = str(shutil.copy(CSV_FILE, tmp_path / "data.csv"))
        lib.load_workbook(path)
        lib.insert_row({"Description": "Only Desc"}, row=2)
        inserted = lib.get_rows()[0]
        assert inserted["Description"] == "Only Desc"
        assert inserted["Product ID"] == ""
        assert inserted["Price"] == ""
        assert inserted["Location"] == ""

    def test_row_is_persisted_after_save(self, lib: RFExcelLibrary, tmp_path: Path):
        path = str(shutil.copy(CSV_FILE, tmp_path / "data.csv"))
        lib.load_workbook(path)
        lib.insert_row(_NEW_ROW, row=2)
        lib.save_workbook()
        lib.close()

        lib2 = RFExcelLibrary()
        lib2.load_workbook(path)
        rows = lib2.get_rows()
        assert rows[0]["Product ID"] == "P-NEW"
        assert rows[0]["Description"] == "Inserted"
        assert rows[0]["Price"] == 5.55
        assert rows[0]["Location"] == "Depot"
        lib2.close()


# ---------------------------------------------------------------------------
# No workbook open
# ---------------------------------------------------------------------------

class TestInsertRowNoWorkbook:

    def test_insert_row_raises_when_no_workbook_open(
        self, lib: RFExcelLibrary
    ):
        with pytest.raises(WorkbookNotOpenException):
            lib.insert_row(_NEW_ROW, row=2)


# ---------------------------------------------------------------------------
# XLSX – Shifted table (headers do NOT start at column A)
# ---------------------------------------------------------------------------

class TestInsertRowXlsxShifted:

    def _make_shifted_xlsx(self, tmp_path: Path) -> str:
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws["B1"] = "Product ID"
        ws["C1"] = "Description"
        ws["D1"] = "Price"
        ws["E1"] = "Location"
        ws["B2"] = "P-001"
        ws["C2"] = "Alpha"
        ws["D2"] = 1.00
        ws["E2"] = "Store"
        ws["B3"] = "P-002"
        ws["C3"] = "Beta"
        ws["D3"] = 2.00
        ws["E3"] = "Warehouse"
        path = str(tmp_path / "shifted.xlsx")
        wb.save(path)
        return path

    def test_new_row_lands_in_correct_columns(self, lib: RFExcelLibrary, tmp_path: Path):
        path = self._make_shifted_xlsx(tmp_path)
        lib.load_workbook(path)
        lib.insert_row(
            {"Product ID": "P-NEW", "Description": "Widget", "Price": 9.99, "Location": "Online"},
            row=2,
        )
        lib.save_workbook()

        wb = openpyxl.load_workbook(path)
        ws = wb.active
        assert ws is not None
        assert ws.cell(2, 1).value is None,    "Column A must stay empty"
        assert ws.cell(2, 2).value == "P-NEW",  "Product ID must land in col B"
        assert ws.cell(2, 3).value == "Widget", "Description must land in col C"
        assert ws.cell(2, 4).value == 9.99,   "Price must land in col D"
        assert ws.cell(2, 5).value == "Online", "Location must land in col E"
        assert ws.cell(3, 2).value == "P-001"
