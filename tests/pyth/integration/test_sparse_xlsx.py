from pathlib import Path

import pytest
from openpyxl import Workbook

from rfexcel.RFExcelLibrary import RFExcelLibrary


def _make_offset_xlsx(path: str, start_col: int = 2) -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.cell(row=1, column=start_col,     value="Name")
    ws.cell(row=1, column=start_col + 1, value="Score")
    ws.cell(row=2, column=start_col,     value="Alice")
    ws.cell(row=2, column=start_col + 1, value=90)
    ws.cell(row=3, column=start_col,     value="Bob")
    ws.cell(row=3, column=start_col + 1, value=75)
    wb.save(path)
    wb.close()


def _make_gap_xlsx(path: str) -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.cell(row=1, column=1, value="Name")
    ws.cell(row=1, column=3, value="Score")
    ws.cell(row=2, column=1, value="Alice")
    ws.cell(row=2, column=3, value=90)
    ws.cell(row=3, column=1, value="Bob")
    ws.cell(row=3, column=3, value=75)
    wb.save(path)
    wb.close()


# ---------------------------------------------------------------------------
# Offset table – shared edit/stream behaviour
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("read_only", [False, True], ids=["xlsx_edit", "xlsx_stream"])
def test_offset_table_correct_row_count(lib: RFExcelLibrary, tmp_path: Path, read_only: bool):
    path = str(tmp_path / "offset.xlsx")
    _make_offset_xlsx(path)
    lib.load_workbook(path, read_only=read_only)
    assert len(lib.get_rows()) == 2


@pytest.mark.parametrize("read_only", [False, True], ids=["xlsx_edit", "xlsx_stream"])
def test_offset_table_values_mapped_to_correct_columns(
    lib: RFExcelLibrary, tmp_path: Path, read_only: bool
):
    path = str(tmp_path / "offset.xlsx")
    _make_offset_xlsx(path)
    lib.load_workbook(path, read_only=read_only)
    rows = lib.get_rows()
    assert rows[0]["Name"] == "Alice"
    assert rows[0]["Score"] == 90
    assert rows[1]["Name"] == "Bob"
    assert rows[1]["Score"] == 75


# ---------------------------------------------------------------------------
# Offset table – edit-only tests
# ---------------------------------------------------------------------------

class TestOffsetTableXlsxEdit:

    def test_header_keys_are_correct(self, lib: RFExcelLibrary, tmp_path: Path):
        path = str(tmp_path / "offset.xlsx")
        _make_offset_xlsx(path)
        lib.load_workbook(path)
        rows = lib.get_rows()
        assert list(rows[0]) == ["Name", "Score"]

    def test_column_c_start(self, lib: RFExcelLibrary, tmp_path: Path):
        path = str(tmp_path / "offset_c.xlsx")
        _make_offset_xlsx(path, start_col=3)
        lib.load_workbook(path)
        rows = lib.get_rows()
        assert rows[0]["Name"] == "Alice"
        assert rows[0]["Score"] == 90


# ---------------------------------------------------------------------------
# Gap column – shared edit/stream behaviour
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("read_only", [False, True], ids=["xlsx_edit", "xlsx_stream"])
def test_gap_column_correct_row_count(lib: RFExcelLibrary, tmp_path: Path, read_only: bool):
    path = str(tmp_path / "gap.xlsx")
    _make_gap_xlsx(path)
    lib.load_workbook(path, read_only=read_only)
    assert len(lib.get_rows()) == 2


@pytest.mark.parametrize("read_only", [False, True], ids=["xlsx_edit", "xlsx_stream"])
def test_gap_column_values_skip_gap_correctly(
    lib: RFExcelLibrary, tmp_path: Path, read_only: bool
):
    path = str(tmp_path / "gap.xlsx")
    _make_gap_xlsx(path)
    lib.load_workbook(path, read_only=read_only)
    rows = lib.get_rows()
    assert rows[0]["Name"] == "Alice"
    assert rows[0]["Score"] == 90
    assert rows[1]["Name"] == "Bob"
    assert rows[1]["Score"] == 75


# ---------------------------------------------------------------------------
# Gap column – edit-only tests
# ---------------------------------------------------------------------------

class TestGapColumnXlsxEdit:

    def test_header_keys_exclude_empty_column(self, lib: RFExcelLibrary, tmp_path: Path):
        path = str(tmp_path / "gap.xlsx")
        _make_gap_xlsx(path)
        lib.load_workbook(path)
        rows = lib.get_rows()
        assert list(rows[0]) == ["Name", "Score"]
