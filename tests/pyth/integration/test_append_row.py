import shutil
from pathlib import Path

import openpyxl
import pytest

from rfexcel.exception.library_exceptions import (
    HeadersNotDeterminedException, NullComponentException,
    WorkbookNotOpenException)
from rfexcel.RFExcelLibrary import RFExcelLibrary
from tests.pyth.conftest import CSV_FILE, XLS_FILE, XLSX_FILE

_HEADERS = ["Product ID", "Description", "Price", "Location"]
_FULL_ROW = {"Product ID": "P-999", "Description": "Widget", "Price": "9.99", "Location": "Online"}


# ---------------------------------------------------------------------------
# XLSX – Edit mode
# ---------------------------------------------------------------------------

class TestAppendRowXlsxEdit:

    def test_full_row_appears_at_end(self, lib: RFExcelLibrary, tmp_path: Path):
        path = str(shutil.copy(XLSX_FILE, tmp_path / "data.xlsx"))
        lib.load_workbook(path)
        before = len(lib.get_rows())
        lib.append_row(_FULL_ROW)
        rows = lib.get_rows()
        assert len(rows) == before + 1
        assert rows[-1]["Product ID"] == "P-999"
        assert rows[-1]["Description"] == "Widget"
        assert rows[-1]["Price"] == "9.99"
        assert rows[-1]["Location"] == "Online"

    def test_partial_row_fills_missing_columns_with_empty_string(
        self, lib: RFExcelLibrary, tmp_path: Path
    ):
        path = str(shutil.copy(XLSX_FILE, tmp_path / "data.xlsx"))
        lib.load_workbook(path)
        lib.append_row({"Product ID": "P-888", "Price": "0.01"})
        last = lib.get_rows()[-1]
        assert last["Product ID"] == "P-888"
        assert last["Price"] == "0.01"
        assert last["Description"] is None
        assert last["Location"] is None

    def test_unknown_keys_are_silently_ignored(self, lib: RFExcelLibrary, tmp_path: Path):
        path = str(shutil.copy(XLSX_FILE, tmp_path / "data.xlsx"))
        lib.load_workbook(path)
        before = len(lib.get_rows())
        lib.append_row({"Product ID": "P-777", "NonExistent": "ignored"})
        rows = lib.get_rows()
        assert len(rows) == before + 1
        assert rows[-1]["Product ID"] == "P-777"

    def test_row_is_persisted_after_save(self, lib: RFExcelLibrary, tmp_path: Path):
        path = str(shutil.copy(XLSX_FILE, tmp_path / "data.xlsx"))
        lib.load_workbook(path)
        lib.append_row(_FULL_ROW)
        lib.save_workbook()
        lib.close()

        lib2 = RFExcelLibrary()
        lib2.load_workbook(path)
        rows = lib2.get_rows()
        assert rows[-1]["Product ID"] == "P-999"
        lib2.close()

    def test_custom_header_row(self, lib: RFExcelLibrary, tmp_path: Path):
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["filler"])
        ws.append(["Name", "Score"])
        ws.append(["Alice", "90"])
        out = str(tmp_path / "custom.xlsx")
        wb.save(out)

        lib.load_workbook(out)
        lib.append_row({"Name": "Bob", "Score": "85"}, header_row=2)
        rows = lib.get_rows(header_row=2)
        assert rows[-1]["Name"] == "Bob"
        assert rows[-1]["Score"] == "85"

    def test_header_row_out_of_range_raises(self, lib: RFExcelLibrary, tmp_path: Path):
        path = str(shutil.copy(XLSX_FILE, tmp_path / "data.xlsx"))
        lib.load_workbook(path)
        with pytest.raises(HeadersNotDeterminedException):
            lib.append_row(_FULL_ROW, header_row=9999)


# ---------------------------------------------------------------------------
# XLSX – Streaming mode
# ---------------------------------------------------------------------------

class TestAppendRowXlsxStream:

    def test_append_row_raises_in_stream_mode(self, lib: RFExcelLibrary, tmp_path: Path):
        path = str(shutil.copy(XLSX_FILE, tmp_path / "data.xlsx"))
        lib.load_workbook(path, read_only=True)
        with pytest.raises(NullComponentException):
            lib.append_row(_FULL_ROW)


# ---------------------------------------------------------------------------
# XLS – Edit mode
# ---------------------------------------------------------------------------

class TestAppendRowXlsEdit:

    def test_append_row_triggers_conversion_and_persists(
        self, lib: RFExcelLibrary, tmp_path: Path
    ):
        path = str(shutil.copy(XLS_FILE, tmp_path / "example.xls"))
        new_path = str(tmp_path / "result.xlsx")
        lib.load_workbook(path)
        lib.append_row({"First Name": "Jane", "Last Name": "Doe"})
        lib.save_workbook(new_path)
        lib.close()

        lib2 = RFExcelLibrary()
        lib2.load_workbook(new_path)
        rows = lib2.get_rows()
        assert rows[-1]["First Name"] == "Jane"
        assert rows[-1]["Last Name"] == "Doe"
        lib2.close()

    def test_original_xls_untouched_after_append_row(
        self, lib: RFExcelLibrary, tmp_path: Path
    ):
        path = str(shutil.copy(XLS_FILE, tmp_path / "example.xls"))
        original_rows_count = RFExcelLibrary()
        original_rows_count.load_workbook(path)
        before = len(original_rows_count.get_rows())
        original_rows_count.close()

        lib.load_workbook(path)
        lib.append_row({"First Name": "Jane", "Last Name": "Doe"})
        lib.save_workbook(str(tmp_path / "out.xlsx"))
        lib.close()

        lib2 = RFExcelLibrary()
        lib2.load_workbook(path)
        assert len(lib2.get_rows()) == before
        lib2.close()


# ---------------------------------------------------------------------------
# XLS – On demand / Streaming mode
# ---------------------------------------------------------------------------

class TestAppendRowXlsStream:

    def test_append_row_raises_in_xls_stream_mode(self, lib: RFExcelLibrary, tmp_path: Path):
        path = str(shutil.copy(XLS_FILE, tmp_path / "example.xls"))
        lib.load_workbook(path, read_only=True)
        with pytest.raises(NullComponentException):
            lib.append_row(_FULL_ROW)


# ---------------------------------------------------------------------------
# CSV – Edit mode
# ---------------------------------------------------------------------------

class TestAppendRowCsvEdit:

    def test_full_row_appears_at_end(self, lib: RFExcelLibrary, tmp_path: Path):
        path = str(shutil.copy(CSV_FILE, tmp_path / "data.csv"))
        lib.load_workbook(path)
        before = len(lib.get_rows())
        lib.append_row(_FULL_ROW)
        rows = lib.get_rows()
        assert len(rows) == before + 1
        assert rows[-1]["Product ID"] == "P-999"

    def test_partial_row_fills_missing_with_empty_string(
        self, lib: RFExcelLibrary, tmp_path: Path
    ):
        path = str(shutil.copy(CSV_FILE, tmp_path / "data.csv"))
        lib.load_workbook(path)
        lib.append_row({"Description": "Only Desc"})
        last = lib.get_rows()[-1]
        assert last["Description"] == "Only Desc"
        assert last["Product ID"] == ""   # explicitly written as "" by append_row
        assert last["Price"] is None      # column beyond written range → absent
        assert last["Location"] is None   # column beyond written range → absent

    def test_row_is_persisted_after_save(self, lib: RFExcelLibrary, tmp_path: Path):
        path = str(shutil.copy(CSV_FILE, tmp_path / "data.csv"))
        lib.load_workbook(path)
        lib.append_row(_FULL_ROW)
        lib.save_workbook()
        lib.close()

        lib2 = RFExcelLibrary()
        lib2.load_workbook(path)
        rows = lib2.get_rows()
        assert rows[-1]["Product ID"] == "P-999"
        assert rows[-1]["Description"] == "Widget"
        lib2.close()

    def test_multiple_rows_added_in_correct_order(
        self, lib: RFExcelLibrary, tmp_path: Path
    ):
        path = str(shutil.copy(CSV_FILE, tmp_path / "data.csv"))
        lib.load_workbook(path)
        lib.append_row({"Product ID": "P-A"})
        lib.append_row({"Product ID": "P-B"})
        rows = lib.get_rows()
        assert rows[-2]["Product ID"] == "P-A"
        assert rows[-1]["Product ID"] == "P-B"


# ---------------------------------------------------------------------------
# CSV – Streaming mode
# ---------------------------------------------------------------------------

class TestAppendRowCsvStream:

    def test_append_row_raises_in_csv_stream_mode(self, lib: RFExcelLibrary, tmp_path: Path):
        path = str(shutil.copy(CSV_FILE, tmp_path / "data.csv"))
        lib.load_workbook(path, read_only=True)
        with pytest.raises(NullComponentException):
            lib.append_row(_FULL_ROW)


# ---------------------------------------------------------------------------
# No workbook open
# ---------------------------------------------------------------------------

class TestAppendRowNoWorkbook:

    def test_append_row_raises_when_no_workbook_open(
        self, lib: RFExcelLibrary
    ):
        with pytest.raises(WorkbookNotOpenException):
            lib.append_row(_FULL_ROW)


# ---------------------------------------------------------------------------
# XLSX – Shifted table (headers do NOT start at column A)
# ---------------------------------------------------------------------------

class TestAppendRowXlsxShifted:
    def _make_shifted_xlsx(self, tmp_path: Path) -> str:
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws["B1"] = "Product ID"
        ws["C1"] = "Description"
        ws["D1"] = "Price"
        ws["E1"] = "Location"
        ws["B2"] = "P-001"
        ws["C2"] = "Alpha"
        ws["D2"] = "1.00"
        ws["E2"] = "Store"
        ws["B3"] = "P-002"
        ws["C3"] = "Beta"
        ws["D3"] = "2.00"
        ws["E3"] = "Warehouse"
        path = str(tmp_path / "shifted.xlsx")
        wb.save(path)
        return path

    def test_new_row_lands_in_correct_columns(self, lib: RFExcelLibrary, tmp_path: Path):
        path = self._make_shifted_xlsx(tmp_path)
        lib.load_workbook(path)
        lib.append_row({"Product ID": "P-999", "Description": "Widget", "Price": "9.99", "Location": "Online"})
        lib.save_workbook()

        wb = openpyxl.load_workbook(path)
        ws = wb.active
        assert ws is not None
        last_row = ws.max_row

        assert ws.cell(last_row, 1).value is None,  "Column A must stay empty"
        assert ws.cell(last_row, 2).value == "P-999",  "Product ID must land in col B"
        assert ws.cell(last_row, 3).value == "Widget",  "Description must land in col C"
        assert ws.cell(last_row, 4).value == "9.99",   "Price must land in col D"
        assert ws.cell(last_row, 5).value == "Online", "Location must land in col E"

    def test_partial_row_leaves_other_columns_empty(self, lib: RFExcelLibrary, tmp_path: Path):
        path = self._make_shifted_xlsx(tmp_path)
        lib.load_workbook(path)
        lib.append_row({"Product ID": "P-777", "Price": "7.77"})
        lib.save_workbook()

        wb = openpyxl.load_workbook(path)
        ws = wb.active
        assert ws is not None
        last_row = ws.max_row

        assert ws.cell(last_row, 1).value is None,    "Column A stays empty"
        assert ws.cell(last_row, 2).value == "P-777", "Product ID lands in col B"
        assert ws.cell(last_row, 3).value is None,    "Description not provided → empty"
        assert ws.cell(last_row, 4).value == "7.77",  "Price lands in col D"
        assert ws.cell(last_row, 5).value is None,    "Location not provided → empty"

    def test_get_rows_still_returns_correct_dict(self, lib: RFExcelLibrary, tmp_path: Path):
        path = self._make_shifted_xlsx(tmp_path)
        lib.load_workbook(path)
        lib.append_row({"Product ID": "P-888", "Description": "Gamma", "Price": "8.88", "Location": "Depot"})
        rows = lib.get_rows()
        assert rows[-1]["Product ID"] == "P-888"
        assert rows[-1]["Description"] == "Gamma"
        assert rows[-1]["Price"] == "8.88"
        assert rows[-1]["Location"] == "Depot"