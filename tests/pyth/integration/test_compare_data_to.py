from pathlib import Path
from typing import cast

import openpyxl
import pytest

from rfexcel.exception.library_exceptions import (NotMatchingColumns,
                                                  WorkbookNotOpenException)
from rfexcel.RFExcelLibrary import RFExcelLibrary
from rfexcel.utils.types import ColumnDifference
from tests.pyth.conftest import CSV_FILE, XLS_FILE, XLSX2_FILE, XLSX_FILE

XLSX_VS_XLSX2_DIFFS = [
    {
        "source_row_index": 5,
        "differences": {
            "Product ID": {"source": "P-203", "target": "P-205"},
            "Price":      {"source": 5.99,  "target": 6},
        },
    },
]

XLSX_VS_CSV_DIFFS = [
    {
        "source_row_index": 2,
        "differences": {
            "Price": {"source": 25.5, "target": "25.50"},
        },
    },
    {
        "source_row_index": 3,
        "differences": {
            "Description": {"source": "Keyboard, Mechanical", "target": "Keyboard, Mechanical, RGB"},
        },
    },
    {
        "source_row_index": 4,
        "differences": {
            "Price": {"source": 150, "target": "150.00"},
        },
    },
    {
        "source_row_index": 5,
        "differences": {
            "Description": {"source": "USB Cable",  "target": "USB Cable, 3ft"},
            "Location":    {"source": "OnlineP",    "target": "Online"},
        },
    },
]

# ---------------------------------------------------------------------------
# xlsx vs xlsx2
# ---------------------------------------------------------------------------

class TestCompareDataToXlsxVsXlsx2:

    def test_returns_list(self, lib: RFExcelLibrary):
        lib.load_workbook(XLSX_FILE)
        assert isinstance(lib.compare_data_to(XLSX2_FILE), list)

    def test_only_row_5_differs(self, lib: RFExcelLibrary):
        lib.load_workbook(XLSX_FILE)
        result = lib.compare_data_to(XLSX2_FILE)
        assert len(result) == 1
        assert result[0]["source_row_index"] == 5

    def test_product_id_diff_on_row_5(self, lib: RFExcelLibrary):
        lib.load_workbook(XLSX_FILE)
        diffs: ColumnDifference  = cast(ColumnDifference, lib.compare_data_to(XLSX2_FILE)[0]["differences"])
        assert diffs["Product ID"] == {"source": "P-203", "target": "P-205"}

    def test_price_diff_on_row_5(self, lib: RFExcelLibrary):
        lib.load_workbook(XLSX_FILE)
        diffs: ColumnDifference  = cast(ColumnDifference, lib.compare_data_to(XLSX2_FILE)[0]["differences"])
        assert diffs["Price"] == {"source": 5.99, "target": 6}

    def test_unchanged_columns_absent_from_differences(self, lib: RFExcelLibrary):
        lib.load_workbook(XLSX_FILE)
        diffs: ColumnDifference  = cast(ColumnDifference, lib.compare_data_to(XLSX2_FILE)[0]["differences"])
        assert "Description" not in diffs
        assert "Location" not in diffs

    def test_identical_rows_not_reported(self, lib: RFExcelLibrary):
        lib.load_workbook(XLSX_FILE)
        result = lib.compare_data_to(XLSX2_FILE)
        reported: set[int] = {entry["source_row_index"] for entry in result}
        assert reported.isdisjoint({2, 3, 4})

    def test_full_result_matches_expected_structure(self, lib: RFExcelLibrary):
        lib.load_workbook(XLSX_FILE)
        assert lib.compare_data_to(XLSX2_FILE) == XLSX_VS_XLSX2_DIFFS

    def test_source_in_stream_mode_same_result(self, lib: RFExcelLibrary):
        lib.load_workbook(XLSX_FILE, read_only=True)
        assert lib.compare_data_to(XLSX2_FILE) == XLSX_VS_XLSX2_DIFFS

    def test_headers_subset_only_reports_requested_columns(self, lib: RFExcelLibrary):
        lib.load_workbook(XLSX_FILE)
        result = lib.compare_data_to(XLSX2_FILE, headers=["Product ID"])
        assert len(result) == 1
        assert "Product ID" in result[0]["differences"]
        assert "Price" not in result[0]["differences"]

    def test_headers_identical_column_returns_no_diffs(self, lib: RFExcelLibrary):
        lib.load_workbook(XLSX_FILE)
        assert lib.compare_data_to(XLSX2_FILE, headers=["Description"]) == []


# ---------------------------------------------------------------------------
# target_sheet parameter
# ---------------------------------------------------------------------------

class TestCompareDataToTargetSheet:

    def test_default_uses_first_sheet(self, lib: RFExcelLibrary):
        lib.load_workbook(XLSX_FILE)
        default_result = lib.compare_data_to(XLSX2_FILE)
        lib.close()
        lib.load_workbook(XLSX_FILE)
        explicit_result = lib.compare_data_to(XLSX2_FILE, target_sheet="List 1")
        assert default_result == explicit_result

    def test_identical_target_sheet_returns_empty(self, lib: RFExcelLibrary):
        lib.load_workbook(XLSX_FILE)
        lib.switch_sheet("Sheet2")
        assert lib.compare_data_to(XLSX2_FILE, target_sheet="Sheet2") == []

    def test_different_source_sheet_still_compares_against_target_sheet(self, lib: RFExcelLibrary):
        lib.load_workbook(XLSX_FILE)
        result = lib.compare_data_to(XLSX2_FILE, target_sheet="Sheet2")
        assert len(result) == 4


# ---------------------------------------------------------------------------
# xlsx vs csv
# ---------------------------------------------------------------------------

class TestCompareDataToXlsxVsCsv:

    def test_returns_four_diff_entries(self, lib: RFExcelLibrary):
        lib.load_workbook(XLSX_FILE)
        assert len(lib.compare_data_to(CSV_FILE)) == 4

    def test_row_3_description_differs(self, lib: RFExcelLibrary):
        lib.load_workbook(XLSX_FILE)
        result = lib.compare_data_to(CSV_FILE)
        row3 = next(e for e in result if e["source_row_index"] == 3)
        assert row3["differences"]["Description"] == {
            "source": "Keyboard, Mechanical",
            "target": "Keyboard, Mechanical, RGB",
        }

    def test_row_5_description_differs(self, lib: RFExcelLibrary):
        lib.load_workbook(XLSX_FILE)
        result = lib.compare_data_to(CSV_FILE)
        row5 = next(e for e in result if e["source_row_index"] == 5)
        assert row5["differences"]["Description"] == {"source": "USB Cable", "target": "USB Cable, 3ft"}

    def test_row_5_location_differs(self, lib: RFExcelLibrary):
        lib.load_workbook(XLSX_FILE)
        result = lib.compare_data_to(CSV_FILE)
        row5 = next(e for e in result if e["source_row_index"] == 5)
        assert row5["differences"]["Location"] == {"source": "OnlineP", "target": "Online"}

    def test_full_result_matches_expected_structure(self, lib: RFExcelLibrary):
        lib.load_workbook(XLSX_FILE)
        assert lib.compare_data_to(CSV_FILE) == XLSX_VS_CSV_DIFFS

    def test_price_differs_where_csv_has_trailing_zeros(self, lib: RFExcelLibrary):
        lib.load_workbook(XLSX_FILE)
        result = lib.compare_data_to(CSV_FILE, headers=["Price"])
        assert len(result) == 2

    def test_product_id_is_identical_across_all_rows(self, lib: RFExcelLibrary):
        lib.load_workbook(XLSX_FILE)
        assert lib.compare_data_to(CSV_FILE, headers=["Product ID"]) == []


# ---------------------------------------------------------------------------
# header row variants
# ---------------------------------------------------------------------------

class TestCompareDataToHeaderRow:

    def test_custom_header_rows_both_offset(self, lib: RFExcelLibrary, tmp_path: Path):
        def _make(path: str, pid: str) -> None:
            wb = openpyxl.Workbook()
            ws = wb.active
            assert ws is not None
            ws.append(["Title"])
            ws.append(["Product ID", "Price"])
            ws.append([pid, "9.99"])
            wb.save(path)

        source_path = str(tmp_path / "source.xlsx")
        target_path = str(tmp_path / "target.xlsx")
        _make(source_path, "P-001")
        _make(target_path, "P-002")

        lib.load_workbook(source_path)
        result = lib.compare_data_to(target_path, source_header_row=2, target_header_row=2)
        assert len(result) == 1
        assert result[0]["differences"]["Product ID"] == {"source": "P-001", "target": "P-002"}

    def test_identical_files_with_offset_header_returns_empty(self, lib: RFExcelLibrary, tmp_path: Path):
        def _make(path: str) -> None:
            wb = openpyxl.Workbook()
            ws = wb.active
            assert ws is not None
            ws.append(["ignore"])
            ws.append(["Name", "Score"])
            ws.append(["Alice", "90"])
            wb.save(path)

        source_path = str(tmp_path / "s.xlsx")
        target_path = str(tmp_path / "t.xlsx")
        _make(source_path)
        _make(target_path)

        lib.load_workbook(source_path)
        assert lib.compare_data_to(target_path, source_header_row=2, target_header_row=2) == []


# ---------------------------------------------------------------------------
# negative / edge cases
# ---------------------------------------------------------------------------

class TestCompareDataToNegative:

    def test_no_workbook_raises(self, lib: RFExcelLibrary):
        with pytest.raises(WorkbookNotOpenException):
            lib.compare_data_to(XLSX2_FILE)

    def test_unknown_header_in_headers_list_raises(self, lib: RFExcelLibrary):
        lib.load_workbook(XLSX_FILE)
        with pytest.raises(NotMatchingColumns):
            lib.compare_data_to(XLSX2_FILE, headers=["Nonexistent"])

    def test_source_header_absent_in_target_raises(self, lib: RFExcelLibrary, tmp_path: Path):
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["Product ID", "Price"])
        ws.append(["P-200", "25.50"])
        target_path = str(tmp_path / "subset.xlsx")
        wb.save(target_path)

        lib.load_workbook(XLSX_FILE)
        with pytest.raises(NotMatchingColumns):
            lib.compare_data_to(target_path)


# ---------------------------------------------------------------------------
# same-workbook (same path) comparisons
# ---------------------------------------------------------------------------

class TestCompareDataToSameWorkbook:
    def test_xlsx_same_path_returns_no_differences(self, lib: RFExcelLibrary):
        lib.load_workbook(XLSX_FILE)
        assert lib.compare_data_to(XLSX_FILE) == []

    def test_workbook_remains_open_and_usable_after_same_path_compare(
        self, lib: RFExcelLibrary
    ):
        lib.load_workbook(XLSX_FILE)
        lib.compare_data_to(XLSX_FILE)
        rows = lib.get_rows()
        assert len(rows) > 0

    def test_csv_same_path_returns_no_differences(self, lib: RFExcelLibrary):
        lib.load_workbook(CSV_FILE)
        assert lib.compare_data_to(CSV_FILE) == []

    def test_xls_same_path_returns_no_differences(self, lib: RFExcelLibrary):
        lib.load_workbook(XLS_FILE)
        assert lib.compare_data_to(XLS_FILE) == []

    def test_subset_headers_same_path_returns_no_differences(
        self, lib: RFExcelLibrary
    ):
        lib.load_workbook(XLSX_FILE)
        assert lib.compare_data_to(XLSX_FILE, headers=["Product ID", "Price"]) == []

    def test_same_path_does_not_report_identical_rows(self, lib: RFExcelLibrary):
        lib.load_workbook(XLSX_FILE)
        result = lib.compare_data_to(XLSX_FILE)
        assert result == []

    def test_same_workbook_with_stream_mode_xlsx_same_sheet(self, lib: RFExcelLibrary):
        lib.load_workbook(XLSX_FILE, read_only=True)
        assert lib.compare_data_to(XLSX_FILE) == []

    def test_same_workbook_with_stream_mode_xlsx_different_sheet(self, lib: RFExcelLibrary):
        lib.load_workbook(XLSX_FILE, read_only=True)
        assert lib.compare_data_to(XLSX_FILE, target_sheet="Sheet2") != []

    def test_same_workbook_with_stream_mode_xls_same_sheet(self, lib: RFExcelLibrary):
        lib.load_workbook(XLS_FILE, read_only=True)
        assert lib.compare_data_to(XLS_FILE) == []

    def test_same_workbook_with_stream_mode_xls_different_sheet(self, lib: RFExcelLibrary):
        lib.load_workbook(XLS_FILE, read_only=True)
        with pytest.raises(NotMatchingColumns):
            lib.compare_data_to(XLS_FILE, target_sheet="Second")

    def test_same_workbook_with_stream_mode_csv(self, lib: RFExcelLibrary):
        lib.load_workbook(CSV_FILE, read_only=True)
        assert lib.compare_data_to(CSV_FILE) == []


# ---------------------------------------------------------------------------
# fail_on_diff parameter
# ---------------------------------------------------------------------------

class TestCompareDataToFailOnDiff:
    def test_no_diff_does_not_raise_when_fail_on_diff_true(
        self, loaded_xlsx: RFExcelLibrary
    ):
        result = loaded_xlsx.compare_data_to(XLSX2_FILE, headers=["Description"], fail_on_diff=True)
        assert result == []

    def test_same_file_does_not_raise_when_fail_on_diff_true(
        self, loaded_xlsx: RFExcelLibrary
    ):
        result = loaded_xlsx.compare_data_to(XLSX_FILE, fail_on_diff=True)
        assert result == []

    def test_raises_assertion_error_on_diff(self, loaded_xlsx: RFExcelLibrary):
        with pytest.raises(AssertionError):
            loaded_xlsx.compare_data_to(XLSX2_FILE, fail_on_diff=True)

    def test_assertion_error_message_contains_source_row_index(
        self, loaded_xlsx: RFExcelLibrary
    ):
        with pytest.raises(AssertionError, match=r"source_row_index 5"):
            loaded_xlsx.compare_data_to(XLSX2_FILE, fail_on_diff=True)

    def test_assertion_error_message_contains_diff_column(
        self, loaded_xlsx: RFExcelLibrary
    ):
        with pytest.raises(AssertionError, match="Product ID|Price"):
            loaded_xlsx.compare_data_to(XLSX2_FILE, fail_on_diff=True)

    def test_raises_at_first_diff_not_last(self, loaded_xlsx: RFExcelLibrary):
        with pytest.raises(AssertionError, match=r"source_row_index 2"):
            loaded_xlsx.compare_data_to(CSV_FILE, fail_on_diff=True)

    def test_csv_target_raises_assertion_error(self, loaded_xlsx: RFExcelLibrary):
        with pytest.raises(AssertionError):
            loaded_xlsx.compare_data_to(CSV_FILE, fail_on_diff=True)

    def test_fail_on_diff_respects_headers_filter_no_raise(
        self, loaded_xlsx: RFExcelLibrary
    ):
        result = loaded_xlsx.compare_data_to(XLSX2_FILE, headers=["Description"], fail_on_diff=True)
        assert result == []

    def test_fail_on_diff_respects_headers_filter_raises(
        self, loaded_xlsx: RFExcelLibrary
    ):
        with pytest.raises(AssertionError):
            loaded_xlsx.compare_data_to(XLSX2_FILE, headers=["Product ID"], fail_on_diff=True)

    def test_default_fail_on_diff_returns_list_not_exception(
        self, loaded_xlsx: RFExcelLibrary
    ):
        result = loaded_xlsx.compare_data_to(XLSX2_FILE)
        assert isinstance(result, list)
        assert len(result) == 1

    def test_explicit_false_returns_full_diff_list(self, loaded_xlsx: RFExcelLibrary):
        result = loaded_xlsx.compare_data_to(CSV_FILE, fail_on_diff=False)
        assert result == XLSX_VS_CSV_DIFFS