from datetime import datetime
from typing import Any, Callable, Literal, TypeAlias, cast

import pytest
import xlrd
import xlrd.sheet
from openpyxl import Workbook
from openpyxl.cell.read_only import EmptyCell

from rfexcel.model.cell_data.i_raw_cell_data import IRawCellData
from rfexcel.model.cell_data.null_raw_cell_data import NullRawCellData
from rfexcel.model.cell_data.xls_raw_cell_data import XlsRawCellData
from rfexcel.model.cell_data.xlsx_raw_cell_data import XlsxRawCellData
from rfexcel.utils.types import NativeType

CellFactory: TypeAlias = Callable[[Any], IRawCellData]

# ---------------------------------------------------------------------------
# Factories - build each backend's cell wrapper from a conceptual value
# ---------------------------------------------------------------------------


def _xlrd_cell(ctype: int, value: Any) -> xlrd.sheet.Cell:
    return xlrd.sheet.Cell(cast(Literal[0, 1, 2, 3, 4, 5, 6], ctype), value)


def _make_xlsx_cell(value: Any) -> IRawCellData:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.cell(row=1, column=1, value=value)
    cell = ws.cell(row=1, column=1)
    data = XlsxRawCellData(cell, "A1")
    wb.close()
    return data


def _make_xls_cell(value: Any) -> IRawCellData:
    if value is None or value == "":
        cell = _xlrd_cell(xlrd.XL_CELL_EMPTY, "")
    elif isinstance(value, bool):
        cell = _xlrd_cell(xlrd.XL_CELL_BOOLEAN, value)
    elif isinstance(value, (int, float)):
        cell = _xlrd_cell(xlrd.XL_CELL_NUMBER, float(value))
    else:
        cell = _xlrd_cell(xlrd.XL_CELL_TEXT, str(value))
    return XlsRawCellData(cell, "A1")


_FACTORIES: list[CellFactory] = [_make_xlsx_cell, _make_xls_cell]
_IDS = ["xlsx", "xls"]


# ---------------------------------------------------------------------------
# Cross-backend: value correctness
# ---------------------------------------------------------------------------


@pytest.mark.parametrize("factory", _FACTORIES, ids=_IDS)
@pytest.mark.parametrize(
    ("raw_value", "expected"),
    [
        ("hello", "hello"),
        ("", ""),
        (None, ""),
    ],
    ids=["string", "empty_string", "none"],
)
def test_get_value_string_and_empty(
    factory: CellFactory, raw_value: Any, expected: NativeType
) -> None:
    assert factory(raw_value).get_value() == expected


@pytest.mark.parametrize("factory", _FACTORIES, ids=_IDS)
@pytest.mark.parametrize(
    ("raw_value", "expected"),
    [(42, 42), (0, 0), (-7, -7)],
    ids=["positive", "zero", "negative"],
)
def test_get_value_integer_returns_int_type(
    factory: CellFactory, raw_value: int, expected: int
) -> None:
    result = factory(raw_value).get_value()
    assert result == expected
    assert type(result) is int, f"Expected int, got {type(result).__name__}({result!r})"


@pytest.mark.parametrize("factory", _FACTORIES, ids=_IDS)
@pytest.mark.parametrize(
    ("raw_value", "expected"),
    [(3.14, 3.14), (-2.5, -2.5)],
    ids=["positive_float", "negative_float"],
)
def test_get_value_float_returns_float_type(
    factory: CellFactory, raw_value: float, expected: float
) -> None:
    result = factory(raw_value).get_value()
    assert result == expected
    assert type(result) is float, f"Expected float, got {type(result).__name__}({result!r})"


@pytest.mark.parametrize("factory", _FACTORIES, ids=_IDS)
@pytest.mark.parametrize(
    ("raw_value", "expected"),
    [(True, True), (False, False)],
    ids=["bool_true", "bool_false"],
)
def test_get_value_boolean_returns_strict_bool(
    factory: CellFactory, raw_value: bool, expected: bool
) -> None:
    result = factory(raw_value).get_value()
    assert result == expected
    assert type(result) is bool, f"Expected bool, got {type(result).__name__}({result!r})"


# ---------------------------------------------------------------------------
# xlsx-specific
# ---------------------------------------------------------------------------


def test_xlsx_empty_cell_sentinel_returns_empty_string() -> None:
    """EmptyCell (openpyxl read-only sentinel) must produce ''."""
    assert XlsxRawCellData(EmptyCell(), "A1").get_value() == ""


def test_xlsx_datetime_value_returned_as_datetime() -> None:
    dt = datetime(2024, 6, 1, 12, 0)
    result = _make_xlsx_cell(dt).get_value()
    assert result == dt
    assert type(result) is datetime


# ---------------------------------------------------------------------------
# xls-specific
# ---------------------------------------------------------------------------


@pytest.mark.parametrize(
    ("ctype", "raw_value"),
    [
        (xlrd.XL_CELL_BLANK, ""),
        (xlrd.XL_CELL_ERROR, 0),
    ],
    ids=["blank", "error"],
)
def test_xls_blank_and_error_cells_return_empty_string(
    ctype: int, raw_value: Any
) -> None:
    cell = _xlrd_cell(ctype, raw_value)
    assert XlsRawCellData(cell, "A1").get_value() == ""


def test_xls_whole_float_collapses_to_int() -> None:
    """xlrd stores all numbers as float; a whole float (e.g. 42.0) must become int."""
    cell = _xlrd_cell(xlrd.XL_CELL_NUMBER, 42.0)
    result = XlsRawCellData(cell, "A1").get_value()
    assert result == 42
    assert type(result) is int


# ---------------------------------------------------------------------------
# null
# ---------------------------------------------------------------------------


def test_null_raw_cell_data_returns_empty_string() -> None:
    result = NullRawCellData().get_value()
    assert result == ""
    assert type(result) is str
