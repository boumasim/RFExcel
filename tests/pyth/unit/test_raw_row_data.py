from typing import Any, Callable, TypeAlias

import pytest
import xlrd
import xlrd.sheet
from openpyxl import Workbook

from rfexcel.model.raw_data.csv_raw_row_data import CsvRawRowData
from rfexcel.model.raw_data.i_raw_row_data import IRawRowData
from rfexcel.model.raw_data.null_raw_row_data import NullRawRowData
from rfexcel.model.raw_data.xls_raw_row_data import XlsRawRowData
from rfexcel.model.raw_data.xlsx_raw_row_data import XlsxRawRowData

RawFactory: TypeAlias = Callable[[list[Any]], IRawRowData]


# ---------------------------------------------------------------------------
# Factories - normalise a conceptual value list to each format's storage type
# ---------------------------------------------------------------------------

def _make_csv(values: list[Any]) -> IRawRowData:
    return CsvRawRowData([str(v) if v is not None else "" for v in values])


def _xlrd_cell(ctype: int, value: Any) -> xlrd.sheet.Cell:
    """Build an xlrd Cell while keeping test intent explicit."""
    return xlrd.sheet.Cell(ctype, value)


def _make_xls(values: list[Any]) -> IRawRowData:
    cells: list[xlrd.sheet.Cell] = []
    for v in values:
        if v is None:
            cells.append(_xlrd_cell(xlrd.XL_CELL_EMPTY, ""))
        elif isinstance(v, bool):
            cells.append(_xlrd_cell(xlrd.XL_CELL_BOOLEAN, v))
        elif isinstance(v, (int, float)):
            cells.append(_xlrd_cell(xlrd.XL_CELL_NUMBER, float(v)))
        else:
            cells.append(_xlrd_cell(xlrd.XL_CELL_TEXT, str(v)))
    return XlsRawRowData(cells)


def _make_xlsx_cell_mode(values: list[Any]) -> IRawRowData:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    for col, value in enumerate(values, start=1):
        ws.cell(row=1, column=col, value=value)
    row_data = XlsxRawRowData(tuple(ws[1]))
    wb.close()
    return row_data


_FACTORIES: list[RawFactory] = [
    _make_csv,
    _make_xls,
    _make_xlsx_cell_mode,
]
_IDS = ["csv", "xls", "xlsx_cell_mode"]


# ---------------------------------------------------------------------------
# other edge cases
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("factory", _FACTORIES, ids=_IDS)
def test_col_out_of_bounds_returns_none(factory: RawFactory) -> None:
    """Column index 0 must never wrap to the last element via negative indexing."""
    row = factory(["first", "second"])
    result = row.get_dict_row_data({"invalid_col": 0, "valid_col": 2})
    assert result["invalid_col"] == ""
    assert result["valid_col"] == "second"


@pytest.mark.parametrize("factory", _FACTORIES, ids=_IDS)
def test_get_header_map_skips_none_or_empty_column(factory: RawFactory) -> None:
    """A blank/None cell in a header row must not produce a phantom key."""
    row = factory(["Name", None, "Age"])
    assert row.get_header_map() == {"Name": 1, "Age": 3}


@pytest.mark.parametrize("factory", _FACTORIES, ids=_IDS)
def test_get_header_map_skips_whitespace_only_column(factory: RawFactory) -> None:
    """A whitespace-only cell must be excluded from the header map."""
    row = factory(["Name", "   ", "Age"])
    assert row.get_header_map() == {"Name": 1, "Age": 3}


def test_csv_header_keys_are_stripped() -> None:
    row = CsvRawRowData(["  Product ID  ", " Description", "Location "])
    assert row.get_header_map() == {
        "Product ID": 1,
        "Description": 2,
        "Location": 3,
    }


@pytest.mark.parametrize("factory", _FACTORIES, ids=_IDS)
def test_header_keys_are_stripped_in_all_backends(factory: RawFactory) -> None:
    row = factory(["  Product ID  ", " Description", "Location "])
    assert row.get_header_map() == {
        "Product ID": 1,
        "Description": 2,
        "Location": 3,
    }


@pytest.mark.parametrize("factory", _FACTORIES, ids=_IDS)
@pytest.mark.parametrize(
    ("data_row", "header_map", "expected"),
    [
        (["x"], {"A": 1, "B": 2, "C": 3}, {"A": "x", "B": "", "C": ""}),
        (["x", "y"], {"A": 1, "B": 2, "C": 3, "D": 4}, {"A": "x", "B": "y", "C": "", "D": ""}),
    ],
)
def test_missing_column_returns_empty_string_when_row_is_sheet_padded(
    factory: RawFactory,
    data_row: list[str],
    header_map: dict[str, int],
    expected: dict[str, str],
) -> None:
    row = factory(data_row)
    assert row.get_dict_row_data(header_map) == expected


@pytest.mark.parametrize("factory", _FACTORIES, ids=_IDS)
@pytest.mark.parametrize(
    ("data_row", "expected"),
    [
        (["", "x", None, ""], ["x"])
    ],
)
def test_list_row_data_does_not_pad_with_trailing_empty_cells(
    factory: RawFactory,
    data_row: list[str],
    expected: list[str],
) -> None:
    row = factory(data_row)
    assert row.get_list_row_data() == expected


def test_null_get_list_row_data_warns_about_row_data(monkeypatch: pytest.MonkeyPatch) -> None:
    messages: list[str] = []
    monkeypatch.setattr("rfexcel.model.raw_data.null_raw_row_data.logger.warn", messages.append)

    result = NullRawRowData().get_list_row_data()

    assert result == []
    assert messages == ["No row data values were returned"]
