"""Parametrized unit tests for rfexcel.utils.utilities."""
from pathlib import Path

import pytest
import xlrd
from openpyxl import Workbook

from rfexcel.utils.utilities import (convert_string_to_dict_row_data,
                                     convert_xls_to_xlsx,
                                     headers_to_header_map,
                                     parse_cell_coordinate, safe_number_cast,
                                     search_in_row)
from tests.pyth.conftest import XLS_FILE

# ---------------------------------------------------------------------------
# safe_number_cast
# ---------------------------------------------------------------------------

@pytest.mark.parametrize(
    ("value", "expected"),
    [
        ("42", 42),
        ("3.14", 3.14),
        ("3.0", 3),
        ("-5", -5),
        ("-3.14", -3.14),
        ("0", 0),
        ("0.0", 0),
        ("  42  ", 42),
        ("  3.14  ", 3.14),
        ("hello", "hello"),
        ("", ""),
        ("   ", ""),
        ("1e5", "1e5"),
        ("3.0.0", "3.0.0"),
        ("123abc", "123abc"),
    ],
)
def test_safe_number_cast_value(value: str, expected: str | int | float) -> None:
    assert safe_number_cast(value) == expected


@pytest.mark.parametrize(
    ("value", "expected_type"),
    [
        ("42", int),

        ("3.14", float),
        
        # conversion of float to int
        ("3.0", int),

        ("hello", str),

        ("", str),
    ],
)
def test_safe_number_cast_type(value: str, expected_type: type) -> None:
    assert type(safe_number_cast(value)) is expected_type

# ---------------------------------------------------------------------------
# search_in_row
# ---------------------------------------------------------------------------

@pytest.mark.parametrize(
    ("source_row", "search_criteria", "partial_match", "expected"),
    [
        ({"name": "Alice"}, {}, False, True),
        ({"name": "Alice"}, {}, True, True),
        ({}, {}, False, True),

        ({"name": "Alice", "age": "30"}, {"name": "Alice"}, False, True),
        ({"name": "Alice", "age": "30"}, {"name": "Alice", "age": "30"}, False, True),

        ({"name": "Alice"}, {"name": "Bob"}, False, False),

        ({"name": "Alice"}, {"city": "NY"}, False, False),
        ({"name": "Alice"}, {"city": "NY"}, True, False),

        ({"name": "Alice"}, {"name": "lic"}, True, True),
        ({"name": "Alice"}, {"name": "Alice"}, True, True),

        ({"name": "Alice"}, {"name": "Bob"}, True, False),

        ({"name": "Alice", "city": "NY"}, {"name": "Alice", "city": "NY"}, False, True),
        ({"name": "Alice", "city": "NY"}, {"name": "Alice", "city": "LA"}, False, False),

        # source row converted to int, not normalized
        ({"count": 1.0}, {"count": "1"}, False, False),
        ({"count": 1.0}, {"count": "1"}, True, True),


        ({"count": 42}, {"count": "42"}, False, True),
        # source_row treated as string, not normalized
        ({"count": 42}, {"count": "2"}, True, True),
        # source_row treated as string, not normalized
        ({"count": 1.0}, {"count": "1.0"}, False, True),
    ],
)
def test_search_in_row(
    source_row: dict,
    search_criteria: dict,
    partial_match: bool,
    expected: bool,
) -> None:
    assert search_in_row(source_row, search_criteria, partial_match) is expected


# ---------------------------------------------------------------------------
# headers_to_header_map
# ---------------------------------------------------------------------------

@pytest.mark.parametrize(
    ("headers", "expected"),
    [
        # list form → sequential 1-based indices
        (["A", "B", "C"], {"A": 1, "B": 2, "C": 3}),
        (["Name"], {"Name": 1}),
        ([], {}),

        # empty-string names excluded
        (["A", "", "C"], {"A": 1, "C": 3}),
        (["", "B", ""], {"B": 2}),
        (["", "", ""], {}),

        # dict form returned as-is
        ({"X": 5, "Y": 10}, {"X": 5, "Y": 10}),
        ({"Name": 1, "Age": 3}, {"Name": 1, "Age": 3}),
        ({}, {}),
    ],
)
def test_headers_to_header_map(headers: object, expected: dict[str, int]) -> None:
    assert headers_to_header_map(headers) == expected


# ---------------------------------------------------------------------------
# convert_string_to_dict_row_data
# ---------------------------------------------------------------------------

@pytest.mark.parametrize(
    ("data", "delimiter", "expected"),
    [
        # basic key=value pairs with default delimiter
        ("animal=cat;person=Ted", ";", {"animal": "cat", "person": "Ted"}),

        # single pair
        ("key=value", ";", {"key": "value"}),

        # value that itself contains = (e.g. a URL)
        ("url=http://a.com/b=1", ";", {"url": "http://a.com/b=1"}),

        # empty value
        ("key=", ";", {"key": ""}),

        # custom delimiter
        ("a=1|b=2", "|", {"a": "1", "b": "2"}),

        # segments without = are silently ignored
        ("a=1;garbage;b=2", ";", {"a": "1", "b": "2"}),

        # key whitespace stripped; value leading space NOT stripped
        ("key= value", ";", {"key": " value"}),

        # empty string → empty dict
        ("", ";", {}),

        # only delimiter → all segments empty/no = → empty dict
        (";;;", ";", {}),

        # dict input returned as a copy (not the same object)
        ({"x": "1", "y": "2"}, ";", {"x": "1", "y": "2"}),
    ],
)
def test_convert_string_to_dict_row_data(
    data: dict | str,
    delimiter: str,
    expected: dict[str, str],
) -> None:
    assert convert_string_to_dict_row_data(data, delimiter=delimiter) == expected


def test_convert_string_to_dict_row_data_dict_input_is_a_copy() -> None:
    """Mutating the returned dict must not affect the original."""
    original: dict[str, str] = {"a": "1"}
    result = convert_string_to_dict_row_data(original)
    result["b"] = "2"
    assert "b" not in original


# ---------------------------------------------------------------------------
# parse_cell_coordinate
# ---------------------------------------------------------------------------

@pytest.mark.parametrize(
    ("coordinate", "zero_based", "expected"),
    [
        # one-based (default)
        ("A1", False, (1, 1)),
        ("B2", False, (2, 2)),
        ("C3", False, (3, 3)),
        ("A10", False, (10, 1)),
        ("Z1", False, (1, 26)),
        ("AA1", False, (1, 27)),    # column 27 = AA

        # zero-based
        ("A1", True, (0, 0)),
        ("B2", True, (1, 1)),
        ("C3", True, (2, 2)),
        ("A10", True, (9, 0)),
        ("Z1", True, (0, 25)),
        ("AA1", True, (0, 26)),
    ],
)
def test_parse_cell_coordinate(
    coordinate: str, zero_based: bool, expected: tuple[int, int]
) -> None:
    assert parse_cell_coordinate(coordinate, zero_based=zero_based) == expected


# ---------------------------------------------------------------------------
# convert_xls_to_xlsx
# ---------------------------------------------------------------------------

def test_convert_xls_to_xlsx_returns_workbook() -> None:
    """Result must be a valid openpyxl Workbook with at least one sheet."""
    wb = convert_xls_to_xlsx(Path(XLS_FILE))
    assert isinstance(wb, Workbook)
    assert len(wb.sheetnames) >= 1


def test_convert_xls_to_xlsx_sheet_count_matches_source() -> None:
    """Sheet count in the result must equal the number of sheets in the source."""
    xls_book = xlrd.open_workbook(XLS_FILE)
    expected_count = xls_book.nsheets
    xls_book.release_resources()

    wb = convert_xls_to_xlsx(Path(XLS_FILE))
    assert len(wb.sheetnames) == expected_count


def test_convert_xls_to_xlsx_sheet_names_match_source() -> None:
    """Sheet names in the result must match those in the source."""
    xls_book = xlrd.open_workbook(XLS_FILE)
    expected_names = xls_book.sheet_names()
    xls_book.release_resources()

    wb = convert_xls_to_xlsx(Path(XLS_FILE))
    assert wb.sheetnames == expected_names
