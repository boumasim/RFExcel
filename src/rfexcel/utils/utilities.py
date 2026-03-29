import re
from pathlib import Path

import xlrd
from openpyxl import Workbook
from openpyxl.utils import coordinate_to_tuple

from rfexcel.utils.types import DictRowData, HeaderMap, HeaderSpec

_NUMBER_REGEX = re.compile(r"^-?\d+(?:\.\d+)?$")

def safe_number_cast(value: str) -> str | int | float:
    """
    Transforms string to either float, int or keeps it as a string.
    """
    cleaned = value.strip()
    
    if _NUMBER_REGEX.match(cleaned):
        num = float(cleaned)
        if num.is_integer():
            return int(num)
        return num
    return cleaned

def search_in_row(source_row: DictRowData, search_criteria: dict[str, str], partial_match: bool) -> bool:
    """Returns True if ALL rules in search_criteria match source_row (AND logic).

    Each key-value pair in search_criteria is one rule. A rule matches when:
    - partial_match=False: the value in source_row equals the criteria value exactly.
    - partial_match=True:  the criteria value is a substring of the row value.

    A key in search_criteria that does not exist in source_row causes an
    immediate False return — the criterion cannot be satisfied.
    Returns True only when every rule in search_criteria produces a match.
    An empty search_criteria always returns True.
    """
    for key, criteria_value in search_criteria.items():
        if key not in source_row:
            return False
        row_value_str = str(source_row[key])
        if partial_match:
            if criteria_value not in row_value_str:
                return False
        else:
            if criteria_value != row_value_str:
                return False
    return True


def headers_to_header_map(headers: HeaderSpec) -> HeaderMap:
    """Normalise a header specifier to a ``HeaderMap`` (``{name: 1-based-column-index}``).

    Accepts two forms:
    - ``list[str]``: treated as sequential columns starting at 1
      (``["A", "B", "C"]`` → ``{"A": 1, "B": 2, "C": 3}``).
    - ``HeaderMap`` (``dict[str, int]``): returned as-is (already canonical).

    Empty-string names are excluded from the result in both cases.
    """
    if isinstance(headers, dict):
        return headers
    return {name: i + 1 for i, name in enumerate(headers) if name}


def convert_string_to_dict_row_data(data: dict[str, str] | str, delimiter: str = ';') -> dict[str, str]:
    """Converts a string like ``animal=cat;person=Ted`` into a dict[str, str].

    Each segment separated by ``delimiter`` must contain ``=``. Everything
    before the first ``=`` is the key; everything after is the value. This
    means values that themselves contain ``=`` (e.g. URLs) are handled
    correctly. Whitespace around keys and values is stripped. Segments
    without ``=`` are silently ignored.
    """
    if isinstance(data, dict):
        return dict(data)
    result: dict[str, str] = {}
    for segment in data.split(delimiter):
        segment = segment.strip()
        if '=' not in segment:
            continue
        key, _, value = segment.partition('=')
        result[key.strip()] = value.strip()
    return result

def convert_xls_to_xlsx(xls_path: Path) -> Workbook:
    """
    Converts an .xls file to a new openpyxl Workbook object.
    """
    xls_book = xlrd.open_workbook(str(xls_path), formatting_info=False)
    try:
        xlsx_book = Workbook()
        if xlsx_book.active:
            xlsx_book.remove(xlsx_book.active)
                
        for sheet_idx in range(xls_book.nsheets):
            xls_sheet = xls_book.sheet_by_index(sheet_idx)
            xlsx_sheet = xlsx_book.create_sheet(title=xls_sheet.name)
                
            for row_idx in range(xls_sheet.nrows):
                for col_idx in range(xls_sheet.ncols):
                    xlsx_sheet.cell(
                        row=row_idx + 1,
                        column=col_idx + 1,
                        value=xls_sheet.cell_value(row_idx, col_idx)
                    )
    finally:
        xls_book.release_resources()
    return xlsx_book

def parse_cell_coordinate(coordinate: str, zero_based: bool = False) -> tuple[int, int]:
    """
    Parses a cell coordinate like "A1" into (row_index, column_index).
    If zero_based is True, returns zero-based indices; otherwise, returns one-based.
    Xlrd is 0-based, openpyxl is 1-based, so this function can be used to convert coordinates accordingly.
    """
    
    row, col = coordinate_to_tuple(coordinate)
    if zero_based:
        return row - 1, col - 1
    return row, col