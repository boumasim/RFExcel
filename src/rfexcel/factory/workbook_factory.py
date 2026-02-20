from pathlib import Path
from typing import Any

import xlrd
from openpyxl import Workbook
from openpyxl.reader import excel

from rfexcel.backend.metadata.xls_metadata import XlsMetadata
from rfexcel.backend.metadata.xlsx_metadata import XlsxMetadata
from rfexcel.backend.reader.csv_edit_reader import CsvEditReader
from rfexcel.backend.reader.csv_stream_reader import CsvStreamReader
from rfexcel.backend.reader.xls_on_demand_reader import XlsOnDemandReader
from rfexcel.backend.reader.xls_standart_reader import XlsStandardReader
from rfexcel.backend.reader.xlsx_edit_reader import XlsxEditReader
from rfexcel.backend.reader.xlsx_stream_reader import XlsxStreamReader
from rfexcel.backend.resource.csv_resource import (CsvEditResource,
                                                   CsvStreamResource)
from rfexcel.backend.resource.xls_resource import (XlsEditResource,
                                                   XlsStreamResource)
from rfexcel.backend.resource.xlsx_resource import (XlsxEditResource,
                                                    XlsxStreamResource)
from rfexcel.backend.style.xls_style import XlsStyle
from rfexcel.backend.style.xlsx_style import XlsxStyle
from rfexcel.backend.writer.csv_writer import CsvWriter
from rfexcel.backend.writer.xlsx_writer import XlsxWriter
from rfexcel.exception.library_exceptions import (
    FileAlreadyExistsException, FileDoesNotExistException,
    FileFormatNotSupportedException)
from rfexcel.RFExcel import RFExcel
from rfexcel.rfexcel_constants import (CSV_SUFFIX, VALID_SUFFIXES, XLS_SUFFIX,
                                       XLSX_SUFFIX)


class WorkbookFactory:

    def create_workbook(self, path: str, **kwargs: Any) -> RFExcel:
        file_path: Path = Path(path)
        extension: str = file_path.suffix.lower()

        if extension not in VALID_SUFFIXES: raise FileFormatNotSupportedException()
        if file_path.exists(): raise FileAlreadyExistsException()

        file_path.parent.mkdir(parents=True, exist_ok=True)

        if extension == XLSX_SUFFIX:
            return self._create_xlsx_edit(path = file_path, **kwargs)
        elif extension == XLS_SUFFIX:
            raise FileFormatNotSupportedException(msg="Use xlsx format for creating and editing excel files.")
        elif extension == CSV_SUFFIX:
            return self._create_csv_edit(path=file_path, **kwargs)
        else:
            raise Exception("Exception in create_workbook occured")


    def load_workbook(self, path: str, read_only: bool = False, **kwargs: Any) -> RFExcel:
        file_path: Path = Path(path)
        extension: str = file_path.suffix.lower()

        if extension not in VALID_SUFFIXES: raise FileFormatNotSupportedException()
        if not file_path.exists(): raise FileDoesNotExistException(file_path.name)

        if extension == XLSX_SUFFIX:
            if read_only:
                return self._load_xlsx_stream(path=file_path, **kwargs)
            else:
                return self._load_xlsx_edit(path=file_path, **kwargs)
        elif extension == XLS_SUFFIX:
            if read_only:
                return self._load_xls_on_demand(path=file_path, **kwargs)
            else:
                return self._load_xls_standard(path=file_path, **kwargs)
        elif extension == CSV_SUFFIX:
            if read_only:
                return self._load_csv_stream(path=file_path, **kwargs)
            else:
                return self._load_csv_edit(path=file_path, **kwargs)
        else:
            raise Exception("Exception in load_workbook occured")

    def _create_xlsx_edit(self, path: Path, **kwargs: Any) -> RFExcel:
        wb: Workbook = Workbook(**kwargs)
        ws = wb.active
        assert ws is not None
        wb.save(filename=path)
        return RFExcel(XlsxWriter(), XlsxEditReader(), XlsxStyle(), XlsxMetadata(), XlsxEditResource(wb=wb))

    def _load_xlsx_stream(self, path: Path, **kwargs: Any) -> RFExcel:
        _data_only: bool = bool(kwargs.get('data_only', False))
        wb: Workbook = excel.load_workbook(filename=path, read_only=True, data_only=_data_only)
        return RFExcel(reader=XlsxStreamReader(), style=XlsxStyle(), metadata=XlsxMetadata(), resource=XlsxStreamResource(wb))

    def _load_xlsx_edit(self, path: Path, **kwargs: Any) -> RFExcel:
        wb: Workbook = excel.load_workbook(filename=path, read_only=False, **kwargs)
        return RFExcel(writer=XlsxWriter(), reader=XlsxEditReader(), style=XlsxStyle(), metadata=XlsxMetadata(), resource=XlsxEditResource(wb))

    def _load_xls_on_demand(self, path: Path, **kwargs: Any) -> RFExcel:
        _formating_info: bool = bool(kwargs.get('formatting_info', True))
        wb: xlrd.Book = xlrd.open_workbook(str(path), on_demand=True, formatting_info=_formating_info, **kwargs)
        return RFExcel(reader=XlsOnDemandReader(), style=XlsStyle(), metadata=XlsMetadata(), resource=XlsStreamResource(wb))

    def _load_xls_standard(self, path: Path, **kwargs: Any) -> RFExcel:
        _formating_info: bool = bool(kwargs.get('formatting_info', True))
        wb: xlrd.Book = xlrd.open_workbook(str(path), on_demand=False, formatting_info=_formating_info, **kwargs)
        return RFExcel(reader=XlsStandardReader(), style=XlsStyle(), metadata=XlsMetadata(), resource=XlsEditResource(wb))
    
    def _load_csv_stream(self, path: Path, **kwargs: Any) -> RFExcel:
        """Opens CSV in read-only streaming mode."""
        resource = CsvStreamResource(path, **kwargs)
        return RFExcel(reader=CsvStreamReader(), resource=resource)

    def _load_csv_edit(self, path: Path, **kwargs: Any) -> RFExcel:
        """Opens CSV in buffered edit mode (Read into memory)."""
        resource = CsvEditResource(path, **kwargs)
        return RFExcel(reader=CsvEditReader(), writer=CsvWriter(), resource=resource)
    
    def _create_csv_edit(self, path: Path, **kwargs: Any) -> RFExcel:
        """Creates an empty CSV file and returns an Edit object."""
        with open(path, mode='w', newline='', encoding='utf-8'):
            pass
        return self._load_csv_edit(path, **kwargs)