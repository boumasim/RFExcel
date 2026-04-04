from collections.abc import Iterator
from pathlib import Path
from typing import Any, override

from openpyxl import Workbook
from openpyxl.cell.read_only import EmptyCell
from openpyxl.chartsheet import Chartsheet
from openpyxl.worksheet.worksheet import Worksheet

from rfexcel.exception.library_exceptions import (FileSaveException,
                                                  LibraryException,
                                                  NotSupportedInReadOnlyMode,
                                                  SheetDoesNotExistException,
                                                  StreamingViolationException)
from rfexcel.model.cell_data.i_raw_cell_data import IRawCellData
from rfexcel.model.cell_data.xlsx_raw_cell_data import XlsxRawCellData
from rfexcel.model.raw_data.i_raw_row_data import IRawRowData
from rfexcel.model.raw_data.xlsx_raw_row_data import XlsxRawRowData
from rfexcel.utils.library_logger import logger
from rfexcel.utils.types import ColumnValues, InsertNativeType
from rfexcel.utils.utilities import parse_cell_coordinate

from .i_resource import IResource


class XlsxEditResource(IResource):
    def __init__(self, wb: Workbook, path: Path):
        super().__init__(path)
        self._wb: Workbook = wb
        self._active_sheet: Worksheet | None = wb.worksheets[0] if wb.worksheets else None

    @property
    @override
    def active_sheets(self) -> Worksheet | Chartsheet | None:
        return self._active_sheet
    
    @property
    @override
    def current_sheet(self) -> str:
        if not self._active_sheet:
            raise LibraryException("No active worksheet")
        return self._active_sheet.title

    @property
    @override
    def last_read_row_index(self) -> int:
        return -1

    @override
    def fetch_row(self, row_index: int, **kwargs: Any) -> IRawRowData:
        if not self._active_sheet:
            raise LibraryException("No active worksheet")
        max_rows = self._active_sheet.max_row or 0
        if row_index > max_rows or row_index < 1:
            raise StopIteration()
        row_values = next(
            self._active_sheet.iter_rows(min_row=row_index, max_row=row_index, values_only=False)
        )
        return XlsxRawRowData(row_values)

    @override
    def fetch_cell(self, cell_name: str, **kwargs: Any) -> IRawCellData:
        if not self._active_sheet:
            raise LibraryException("No active worksheet")
        row_index, col_index = parse_cell_coordinate(cell_name)
        return XlsxRawCellData(self._active_sheet.cell(row=row_index, column=col_index), cell_name)

    @override
    def get_sheet_names(self) -> list[str]:
        return list(self._wb.sheetnames)

    @override
    def switch_sheet(self, name: str) -> None:
        if name not in self._wb.sheetnames:
            raise SheetDoesNotExistException(name)
        self._active_sheet = self._wb[name]

    @override
    def add_sheet(self, name: str) -> None:
        ws: Worksheet = self._wb.create_sheet(title=name)
        self._active_sheet = ws

    @override
    def delete_sheet(self, name: str) -> None:
        if name not in self._wb.sheetnames:
            raise SheetDoesNotExistException(name)
        del self._wb[name]
        self._active_sheet = self._wb.worksheets[0] if self._wb.worksheets else None

    @override
    def save(self, path: Path | None = None) -> None:
        target = path or self._path
        if target.suffix.lower() == '.xls':
            logger.warn(
                f"Saving xlsx content to '{target.name}' with a .xls extension. "
                "Consider providing a .xlsx path."
            )
        try:
            self._wb.save(filename=target)
        except Exception as e:
            raise FileSaveException(str(target), str(e)) from e
        self._path = target
        logger.info(f"Workbook saved to '{target.name}'.")

    @override
    def append_row(self, cell_data: ColumnValues) -> None:
        if not self._active_sheet:
            raise LibraryException("No active worksheet")
        next_row = (self._active_sheet.max_row or 0) + 1
        for col, value in cell_data.items():
            self._active_sheet.cell(row=next_row, column=col, value=value)

    @override
    def update_row(self, row_index: int, cell_data: ColumnValues) -> None:
        if not self._active_sheet:
            raise LibraryException("No active worksheet")
        for col, value in cell_data.items():
            self._active_sheet.cell(row=row_index, column=col, value=value)

    @override
    def delete_row(self, row_index: int) -> None:
        if not self._active_sheet:
            raise LibraryException("No active worksheet")
        self._active_sheet.delete_rows(row_index, 1)

    @override
    def insert_row(self, row_index: int, cell_data: ColumnValues) -> None:
        if not self._active_sheet:
            raise LibraryException("No active worksheet")
        self._active_sheet.insert_rows(row_index)
        for col, value in cell_data.items():
            self._active_sheet.cell(row=row_index, column=col, value=value)

    @override
    def set_cell(self, cell_name: str, value: InsertNativeType) -> None:
        if not self._active_sheet:
            raise LibraryException("No active worksheet")
        row_index, col_index = parse_cell_coordinate(cell_name)
        self._active_sheet.cell(row=row_index, column=col_index, value=value)

    @override
    def close(self):
        self._wb.close()


class XlsxStreamResource(IResource):
    def __init__(self, wb: Workbook, path: Path):
        super().__init__(path)
        self._wb = wb
        self._active_sheet = self._wb.worksheets[0] if self._wb.worksheets else None
        self._row_generator: Iterator[tuple[Any, ...]] | None = None
        self._last_read_row_index = 0

    def _get_generator(self) -> Iterator[tuple[Any, ...]]  :
        if self._row_generator is None:
            self._row_generator = (
                self._active_sheet.iter_rows(values_only=False)
                if self._active_sheet
                else iter([])
            )
        return self._row_generator

    @property
    @override
    def active_sheets(self) -> Worksheet | Chartsheet | None:
        return self._active_sheet
    
    @property
    @override
    def current_sheet(self) -> str:
        if not self._active_sheet:
            raise LibraryException("No active worksheet")
        return self._active_sheet.title
    
    @property
    @override
    def last_read_row_index(self) -> int:
        return self._last_read_row_index
    
    @override
    def fetch_row(self, row_index: int, **kwargs: Any) -> IRawRowData:
        gen = self._get_generator()
        while(self._last_read_row_index < row_index - 1):
            next(gen)
            self._last_read_row_index += 1
        row_data = next(gen)
        self._last_read_row_index += 1
        return XlsxRawRowData(row_data)

    @override
    def fetch_cell(self, cell_name: str, **kwargs: Any) -> IRawCellData:
        if not self._active_sheet:
            raise LibraryException("No active worksheet")
        row_index, col_index = parse_cell_coordinate(cell_name)
        if row_index <= self._last_read_row_index:
            raise StreamingViolationException(row_index=row_index, last_read=self._last_read_row_index)
        gen = self._get_generator()
        while self._last_read_row_index < row_index - 1:
            try:
                next(gen)
            except StopIteration:
                return XlsxRawCellData(EmptyCell(), cell_name)
            self._last_read_row_index += 1
        try:
            row_tuple = next(gen)
        except StopIteration:
            return XlsxRawCellData(EmptyCell(), cell_name)
        self._last_read_row_index += 1
        if col_index - 1 >= len(row_tuple):
            return XlsxRawCellData(EmptyCell(), cell_name)
        return XlsxRawCellData(row_tuple[col_index - 1], cell_name)

    @override
    def get_sheet_names(self) -> list[str]:
        return list(self._wb.sheetnames)

    @override
    def switch_sheet(self, name: str) -> None:
        if name not in self._wb.sheetnames:
            raise SheetDoesNotExistException(name)
        if self._row_generator is not None:
            self._row_generator.close()  # type: ignore[attr-defined]
        self._active_sheet = self._wb[name]
        self._row_generator = None
        self._last_read_row_index = 0

    @override
    def add_sheet(self, name: str) -> None:
        raise NotSupportedInReadOnlyMode("Adding sheets is not supported in streaming mode")

    @override
    def delete_sheet(self, name: str) -> None:
        raise NotSupportedInReadOnlyMode("Deleting sheets is not supported in streaming mode")

    @override
    def append_row(self, cell_data: ColumnValues) -> None:
        raise NotSupportedInReadOnlyMode("Appending rows is not supported in streaming mode")

    @override
    def update_row(self, row_index: int, cell_data: ColumnValues) -> None:
        raise NotSupportedInReadOnlyMode("Updating rows is not supported in streaming mode")

    @override
    def delete_row(self, row_index: int) -> None:
        raise NotSupportedInReadOnlyMode("Deleting rows is not supported in streaming mode")

    @override
    def insert_row(self, row_index: int, cell_data: ColumnValues) -> None:
        raise NotSupportedInReadOnlyMode("Inserting rows is not supported in streaming mode")

    @override
    def set_cell(self, cell_name: str, value: InsertNativeType) -> None:
        raise NotSupportedInReadOnlyMode("Set Cell is not supported in streaming mode")

    @override
    def save(self, path: Path | None = None) -> None:
        raise NotSupportedInReadOnlyMode("Saving is not supported in streaming (read-only) mode")

    @override
    def close(self):
        if self._row_generator is not None:
            self._row_generator.close()  # type: ignore[attr-defined]
            self._row_generator = None
        self._wb.close()